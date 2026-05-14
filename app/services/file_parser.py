import base64
import io
import re


def parse_file_content(file_info):
    """根据 MIME 类型解析文件，返回提取的文本"""
    name = file_info.get('name', 'file')
    mime = file_info.get('type', '')
    data_url = file_info.get('dataUrl', '')

    match = re.match(r'data:[^;]*;base64,(.+)', data_url, re.DOTALL)
    if not match:
        return None
    raw = base64.b64decode(match.group(1))

    # 纯文本类型
    text_types = ['text/', 'application/json', 'application/javascript',
                  'application/xml', 'application/x-httpd-php']
    if any(mime.startswith(t) for t in text_types):
        try:
            return raw.decode('utf-8')
        except UnicodeDecodeError:
            import chardet
            enc = chardet.detect(raw)['encoding'] or 'gbk'
            return raw.decode(enc, errors='replace')

    # .docx
    if mime.endswith('officedocument.wordprocessingml.document') or name.endswith('.docx'):
        import docx
        doc = docx.Document(io.BytesIO(raw))
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        return '\n\n'.join(paragraphs)

    # .pdf
    if mime == 'application/pdf' or name.endswith('.pdf'):
        import pypdf
        reader = pypdf.PdfReader(io.BytesIO(raw))
        pages = []
        for page in reader.pages:
            text = page.extract_text()
            if text:
                pages.append(text)
        return '\n\n'.join(pages)

    # .xlsx
    if 'spreadsheet' in mime or name.endswith('.xlsx'):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(raw))
        result = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            result.append(f'[Sheet: {sheet_name}]')
            for row in ws.iter_rows(values_only=True):
                row_str = '\t'.join(str(c) if c is not None else '' for c in row)
                if row_str.strip():
                    result.append(row_str)
        return '\n'.join(result)

    # 未知类型 → 尝试文本读取
    try:
        return raw.decode('utf-8')
    except UnicodeDecodeError:
        return None
